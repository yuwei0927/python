import requests
import threading
import time
import csv
import openpyxl


class Spider():
    HEADERS = {
        'User-Agent': '自定义User-Agent ',
        'referer': 'https://xueqiu.com/hq'
    }

    CK = {
        'Cookie': 'bid=e359ea13a1a714a8dde009786a79be42_j28xt6yf; s=el122u6c2c; device_id=c55491892571a32018d91d86d06dbb0b; aliyungf_tc=AQAAAJompTrlbwcAqO9rZcPwcHL+YLXK; xq_a_token=0a52c567442f1fdd8b09c27e0abb26438e274a7e; xq_a_token.sig=dR_XY4cJjuYM6ujKxH735NKcOpw; xq_r_token=43c6fed2d6b5cc8bc38cc9694c6c1cf121d38471; xq_r_token.sig=8d4jOYdZXEWqSBXOB9N5KuMMZq8; u=931499846343026; __utmt=1; __utma=1.406116928.1493813130.1499846347.1499856796.6; __utmb=1.2.10.1499856796; __utmc=1; __utmz=1.1493813130.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); Hm_lvt_1db88642e346389874251b5a1eded6e3=1499762943,1499763039,1499846343,1499856789; Hm_lpvt_1db88642e346389874251b5a1eded6e3=1499856813'}

    def __init__(self, url, page, thread, export_type):
        self.url = url
        self.page = page
        self.thread = thread
        self.export_type = export_type
        self.urls = []
        self.data = []
        self.export = {
            'txt': self._export_txt,
            'csv': self._export_csv,
            'excel': self._export_excel
        }

    def start(self):
        self._obtain_url()
        self._set_up_thread()
        export_func = self.export.get(self.export_type)
        export_func()

    def _obtain_url(self):
        for i in range(1, self.page + 1):
            self.urls.append(self.url.replace('page=1', 'page={}'.format(i)))

    def _obtain_data(self):
        for url in self.urls:
            r = requests.get(url, headers=self.HEADERS, cookies=self.CK)
            html = r.json()
            data = html.get('stocks')
            for d in data:
                code = d.get('code')
                name = d.get('name')
                current = d.get('current')
                change = d.get('change')
                percent = d.get('percent')
                s = (code, name, current, change, percent)
                self.data.append(s)

    def _set_up_thread(self):
        threads = []
        for i in range(self.thread):
            t = threading.Thread(target=self._obtain_data)
            threads.append(t)
        for t in threads:
            t.start()

        for t in threads:
            t.join()

    def _export_txt(self):
        with open('text1.txt', 'w', encoding='utf8') as f:
            for d in self.data:
                row = ','.join([str(x) for x in d])
                f.write('{}\n'.format(row))


    def _export_csv(self):
        header = ['股票代码', '股票名称', '当前价格', '涨跌额度', '涨跌幅度']
        with open('data1.csv', 'w', encoding='utf8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(self.data)

    def _export_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '股票数据'
        ws['A1'] = '股票代码'
        ws['B1'] = '股票名称'
        ws['C1'] = '当前价格'
        ws['D1'] = '涨跌额度'
        ws['E1'] = '涨跌幅度'
        for i in range(len(self.data)):
            ws.cell(row=i + 2, column=1, value=self.data[i][0])
            ws.cell(row=i + 2, column=2, value=self.data[i][1])
            ws.cell(row=i + 2, column=3, value=self.data[i][2])
            ws.cell(row=i + 2, column=4, value=self.data[i][3])
            ws.cell(row=i + 2, column=5, value=self.data[i][4])
        wb.save('numerical.xlsx')


if __name__ == '__main__':
    print('主程序执行时间为{}'.format(time.ctime()))
    spider = Spider(
        'https://xueqiu.com/stock/cata/stocklist.json?page=1&size=30&order=desc&orderby=percent&type=11%2C12&_=1499856813132',
        3, 3, 'excel')
    spider.start()
    print('主程序结束时间为{}'.format(time.ctime()))
