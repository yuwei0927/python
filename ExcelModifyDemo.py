from openpyxl import Workbook
import os
import xlrd
import datetime

def GetExcelData(listname)#����ͬ������Ŀ�Ľ�����浽��Ӧ�Ļ��ܱ����
	listCount = len(listname)
	dataList = [] #����������һSheet�����������
	for i in range(1, listCount)
		data = xlrd.open_workbook(listname[i])
		table = data.sheets()[0]
		if i == 1:
			dataList.append(table.row_values(0))

		list1 = table.row_table(1)
		list1[0] = (xlrd.xldate.xldate_as_datetime(table.cell(1,0).value, 0).date())#���Ϊ���ڣ�ת�������ڸ�ʽ
		list1[1] = (xlrd.xldate.xldate_as_datetime(table.cell(1,1).value, 0).time())#���Ϊʱ�䣬ת����ʱ���ʽ
		dataList.append(list1)

	return dataList

def RemoveFile(listname)
	if os.path.exists(listname[0]):
		os.remove(listname[0])

if __name__ == "__main__":

	filePath = os.getcwd()
	resultname = 'Result.xlsx'
	RB = 'Rub&Buzz'
	THD = 'THD'
	FR = 'Freqresp'

	RBList = [RB]#��������ͬһ������Ŀ��������Ա�����������е�һ��Ϊ��������
	FRList = [FR]
	THDList = [THD]

	RemoveFile(RBList)
	RemoveFile(THDList)
	RemoveFile(FRList)

	fileList = os.listdir(filePath)
	fileCount = len(fileList)

	TempList = fileList[:]#����һ��ͬ�����б�

	#�Ƴ���Excel�����ļ�
	for i in range(fileCount):
		fileExt = os.path.splitext(os.listdir()[i])[1].split('.')[-1]#��ȡ���ļ�����չ��
		if fileExt not in ['xlsx']:
			TempList.remove(fileList[i])#���Ǳ���ļ����б���ɾ��

	fileList = TempList
	fileCount = len(fileList)

	#�����б��������ֽ��з������
	for i in range(fileCount):
		fileFirstName = os.path.splitext(fileList[i])[0].split(' ')[0]
		if fileFirstName  == RB:
			RBList.append(fileList[i])
		elif fileFirstName  == THD:
			THDList.append(fileList[i])
		elif fileFirstName  == FR:
			FRList.append(fileList[i])

	RBdataList = GetExcelData(RBList)
	THDdataList = GetExcelData(THDList)
	FRdataList = GetExcelData(FRList)

	wb = Workbook()
	Rbws = wb.active
	RBws.titl = RB

	for row in RBDataList:
		Rbwb.append(row)

	Frws = wb.create_sheet(FR)
	for row in FRDataList:
		Frwb.append(row)

	THDws = wb.create_sheet(THD)
	for row in THDDataList:
		THDwb.append(row)

	wb.save(resultname)

	print('END!')